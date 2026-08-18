import os
import re
import secrets
import time
from datetime import datetime, timedelta, timezone
import hashlib
import hmac
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import json
import uuid
import urllib.request
import urllib.error
from urllib.parse import parse_qsl, unquote, urlencode, urlparse
from zoneinfo import ZoneInfo

import qrcode
from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, Response, UploadFile
from fastapi.responses import (
    FileResponse,
    HTMLResponse,
    JSONResponse,
    PlainTextResponse,
    RedirectResponse,
    StreamingResponse,
)
from starlette.exceptions import HTTPException as StarletteHTTPException
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field
from starlette.datastructures import UploadFile as StarletteUploadFile

from ticketbot.database import Database, STATUS_PENDING, normalize_maps_url

BASE_DIR = Path(__file__).resolve().parent
WEB_DIR = BASE_DIR / "miniapp"

load_dotenv()
DATABASE_PATH = os.getenv("DATABASE_PATH", "data/bot.db")
ADMIN_IDS = {int(x) for x in os.getenv("ADMIN_IDS", "").split(",") if x.strip().isdigit()}
BOT_TOKEN = os.getenv("BOT_TOKEN", "").strip()
WEB_APP_URL = os.getenv("WEB_APP_URL", "").rstrip("/")
ADMIN_WEB_PASSWORD = os.getenv("ADMIN_WEB_PASSWORD", "")
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "").strip()
DEFAULT_UPLOAD_DIR = str(Path(DATABASE_PATH).resolve().parent / "uploads")
UPLOAD_DIR = os.getenv("UPLOAD_DIR", DEFAULT_UPLOAD_DIR)
# Event banners are PUBLIC + PERMANENT (unlike signed/expiring payment proofs), so
# they live in a dedicated directory on the Railway /data volume and are served
# unsigned via GET /event-media/{filename}.
DEFAULT_EVENT_MEDIA_DIR = str(Path(DATABASE_PATH).resolve().parent / "event_media")
EVENT_MEDIA_DIR = os.getenv("EVENT_MEDIA_DIR", DEFAULT_EVENT_MEDIA_DIR)
os.makedirs(EVENT_MEDIA_DIR, exist_ok=True)
ALLOW_TG_ID_FALLBACK = os.getenv("MINIAPP_ALLOW_TG_ID_FALLBACK", "0").strip().lower() in {
    "1",
    "true",
    "yes",
}
# REQUIRE_SECURE_CONFIG turns on production hardening. It is OFF by default so the
# development/test fallbacks (dev signing secrets, tg_id query fallback) keep working.
# ENVIRONMENT="production" implicitly enables it.
REQUIRE_SECURE_CONFIG = (
    os.getenv("REQUIRE_SECURE_CONFIG", "0").strip().lower() in {"1", "true", "yes"}
    or os.getenv("ENVIRONMENT", "").strip().lower() == "production"
)


def _enforce_secure_config() -> None:
    """Fail fast at import time when production hardening is required but unsafe.

    When REQUIRE_SECURE_CONFIG is ON we refuse to start with the forgeable tg_id query
    fallback enabled, and we require real signing secrets (dedicated env vars, or a
    BOT_TOKEN to back them) so the getters can never fall back to the 'dev-*' literals.
    """
    if not REQUIRE_SECURE_CONFIG:
        return
    if ALLOW_TG_ID_FALLBACK:
        raise RuntimeError(
            "REQUIRE_SECURE_CONFIG is enabled but MINIAPP_ALLOW_TG_ID_FALLBACK is on. "
            "The tg_id query fallback is forgeable and must be disabled in production."
        )
    if not BOT_TOKEN:
        missing = [
            name
            for name in ("UPLOAD_LINK_SECRET", "TICKET_QR_SECRET", "EMAIL_LOGIN_SECRET")
            if not (os.getenv(name, "") or "").strip()
        ]
        if missing:
            raise RuntimeError(
                "REQUIRE_SECURE_CONFIG is enabled but required signing secrets are missing: "
                + ", ".join(missing)
                + ". Set dedicated secrets (or a BOT_TOKEN to back them) before starting."
            )


_enforce_secure_config()


def _env_positive_float(name: str, default: float) -> float:
    raw = (os.getenv(name, "") or "").strip()
    if not raw:
        return default
    try:
        value = float(raw)
    except ValueError:
        return default
    return value if value > 0 else default


def _env_positive_int(name: str, default: int) -> int:
    raw = (os.getenv(name, "") or "").strip()
    if not raw:
        return default
    try:
        value = int(raw)
    except ValueError:
        return default
    return value if value > 0 else default


UPLOAD_MAX_MB = _env_positive_float("UPLOAD_MAX_MB", 5.0)
MAX_UPLOAD_BYTES = int(UPLOAD_MAX_MB * 1024 * 1024)
TELEGRAM_AUTH_MAX_AGE_SECONDS = _env_positive_int("TELEGRAM_AUTH_MAX_AGE_SECONDS", 86400)
UPLOAD_RETENTION_DAYS = _env_positive_int("UPLOAD_RETENTION_DAYS", 7)
UPLOAD_CLEANUP_INTERVAL_SECONDS = _env_positive_int("UPLOAD_CLEANUP_INTERVAL_SECONDS", 3600)
UPLOAD_LINK_TTL_SECONDS = _env_positive_int("UPLOAD_LINK_TTL_SECONDS", UPLOAD_RETENTION_DAYS * 24 * 60 * 60)
RATE_LIMIT_WINDOW_SECONDS = _env_positive_int("RATE_LIMIT_WINDOW_SECONDS", 60)
QUOTE_RATE_LIMIT = _env_positive_int("QUOTE_RATE_LIMIT", 120)
BOOKING_RATE_LIMIT = _env_positive_int("BOOKING_RATE_LIMIT", 12)
EMAIL_LOGIN_TTL_SECONDS = _env_positive_int("EMAIL_LOGIN_TTL_SECONDS", 600)
EMAIL_LOGIN_RATE_LIMIT = _env_positive_int("EMAIL_LOGIN_RATE_LIMIT", 8)
EMAIL_CODE_ATTEMPT_LIMIT = _env_positive_int("EMAIL_CODE_ATTEMPT_LIMIT", 5)
RESEND_API_KEY = os.getenv("RESEND_API_KEY", "").strip()
RESEND_FROM_EMAIL = os.getenv("RESEND_FROM_EMAIL", "").strip()
EMAIL_LOGIN_DEV_MODE = os.getenv("EMAIL_LOGIN_DEV_MODE", "0").strip().lower() in {"1", "true", "yes"}
LEGACY_WEB_REGISTER_ENABLED = os.getenv("LEGACY_WEB_REGISTER_ENABLED", "0").strip().lower() in {
    "1",
    "true",
    "yes",
}
WEB_SESSION_COOKIE = "bt_web_session"
ADMIN_SESSION_COOKIE = "bt_admin_session"
SESSION_COOKIE_MAX_AGE_SECONDS = _env_positive_int("SESSION_COOKIE_MAX_AGE_SECONDS", 60 * 60 * 24 * 90)
_LAST_UPLOAD_CLEANUP_TS = 0.0
_RATE_LIMIT_BUCKETS: Dict[Tuple[str, str], List[float]] = {}

os.makedirs(UPLOAD_DIR, exist_ok=True)
db = Database(DATABASE_PATH)

app = FastAPI(title="TicketBot Mini App Server")
app.mount("/static", StaticFiles(directory=str(WEB_DIR)), name="static")


@app.middleware("http")
async def add_security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(self), microphone=(), geolocation=()")
    response.headers.setdefault(
        "Content-Security-Policy",
        "default-src 'self'; "
        "script-src 'self' https://telegram.org https://accounts.google.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: https:; "
        "connect-src 'self'; "
        "frame-src https://accounts.google.com https://www.google.com https://maps.google.com; "
        "frame-ancestors https://web.telegram.org https://*.telegram.org;",
    )
    path = request.url.path
    # Keep the intentional no-store policy for HTML/JS/CSS so code deploys take
    # effect immediately. The .js/.css checks are ordered first so they always
    # win no-store, even though they also live under /static.
    if path in {"/", "/admin"} or path.endswith((".html", ".js", ".css")):
        response.headers["Cache-Control"] = "no-store, max-age=0"
    elif path.endswith((".jpg", ".jpeg", ".png", ".svg", ".woff", ".woff2", ".ico")):
        # Static images/fonts are content-addressed via query-version strings,
        # so they can be cached far-future and immutable.
        response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
    return response


def _fallback_auth_allowed() -> bool:
    return ALLOW_TG_ID_FALLBACK


def _parse_telegram_init_data(init_data: str) -> Dict[str, str]:
    return dict(parse_qsl(init_data, keep_blank_values=True, strict_parsing=False))


def _verify_telegram_init_data(init_data: str) -> Dict[str, Any]:
    if not BOT_TOKEN:
        raise HTTPException(status_code=503, detail="Telegram auth is not configured.")
    if not init_data:
        raise HTTPException(status_code=401, detail="Open this Mini App from Telegram.")

    parsed = _parse_telegram_init_data(init_data)
    received_hash = parsed.pop("hash", "")
    if not received_hash:
        raise HTTPException(status_code=401, detail="Telegram auth hash is missing.")

    data_check_string = "\n".join(f"{key}={value}" for key, value in sorted(parsed.items()))
    secret_key = hmac.new(b"WebAppData", BOT_TOKEN.encode("utf-8"), hashlib.sha256).digest()
    expected_hash = hmac.new(secret_key, data_check_string.encode("utf-8"), hashlib.sha256).hexdigest()
    if not hmac.compare_digest(expected_hash, received_hash):
        raise HTTPException(status_code=401, detail="Telegram auth is invalid.")

    try:
        auth_date = int(parsed.get("auth_date", "0"))
    except ValueError as exc:
        raise HTTPException(status_code=401, detail="Telegram auth date is invalid.") from exc
    if TELEGRAM_AUTH_MAX_AGE_SECONDS > 0 and time.time() - auth_date > TELEGRAM_AUTH_MAX_AGE_SECONDS:
        raise HTTPException(status_code=401, detail="Telegram session expired. Reopen the Mini App.")

    try:
        user = json.loads(parsed.get("user", "{}"))
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=401, detail="Telegram user data is invalid.") from exc
    if not isinstance(user, dict) or not user.get("id"):
        raise HTTPException(status_code=401, detail="Telegram user is missing.")
    return user


def _request_tg_id(request: Request, provided_tg_id: Optional[int]) -> int:
    init_data = request.headers.get("x-telegram-init-data", "")
    if init_data:
        user = _verify_telegram_init_data(init_data)
        verified_tg_id = int(user["id"])
        if provided_tg_id is not None and int(provided_tg_id) != verified_tg_id:
            raise HTTPException(status_code=403, detail="Telegram user mismatch.")
        return verified_tg_id
    if provided_tg_id is not None and _fallback_auth_allowed():
        return int(provided_tg_id)
    raise HTTPException(status_code=401, detail="Open this Mini App from Telegram.")


def _request_web_token(request: Request) -> str:
    return (request.cookies.get(WEB_SESSION_COOKIE) or "").strip()


def _request_admin_token(request: Request) -> str:
    return (request.cookies.get(ADMIN_SESSION_COOKIE) or "").strip()


def _cookie_secure(request: Request) -> bool:
    forwarded_proto = request.headers.get("x-forwarded-proto", "").split(",", 1)[0].strip().lower()
    return forwarded_proto == "https" or request.url.scheme == "https"


def _set_session_cookie(request: Request, response: Response, name: str, token: str) -> None:
    response.set_cookie(
        key=name,
        value=token,
        max_age=SESSION_COOKIE_MAX_AGE_SECONDS,
        httponly=True,
        secure=_cookie_secure(request),
        samesite="lax",
        path="/",
    )


def _ensure_not_blocked(user: Any) -> None:
    tg_id = getattr(user, "tg_id", None)
    if getattr(user, "blocked", 0) or (tg_id is not None and db.is_blocked(tg_id)):
        raise HTTPException(
            status_code=403,
            detail="Your account is blocked. Contact the organizers.",
        )


def _request_user(request: Request, provided_tg_id: Optional[int]) -> Tuple[Any, Optional[int]]:
    try:
        tg_id = _request_tg_id(request, provided_tg_id)
    except HTTPException as exc:
        if exc.status_code not in {401, 403}:
            raise
    else:
        user = db.get_user(tg_id)
        if not user:
            raise HTTPException(status_code=404, detail="Create your profile before booking.")
        _ensure_not_blocked(user)
        return user, tg_id

    user = db.get_user_by_web_session(_request_web_token(request))
    if not user:
        raise HTTPException(status_code=401, detail="Register or log in before booking.")
    _ensure_not_blocked(user)
    return user, None


def _extract_upload_filename(payment_file_id: str) -> Optional[str]:
    raw = (payment_file_id or "").strip()
    if not raw:
        return None
    parsed = urlparse(raw)
    path = parsed.path if parsed.scheme else raw
    if not path.startswith("/uploads/"):
        return None
    filename = unquote(Path(path).name)
    if not filename or filename in {".", ".."}:
        return None
    return filename


def _delete_stored_upload(upload_url: str) -> None:
    filename = _extract_upload_filename(upload_url)
    if not filename:
        return
    try:
        _safe_upload_path(filename).unlink()
    except HTTPException:
        return
    except OSError:
        return


def _upload_signing_secret() -> bytes:
    secret = os.getenv("UPLOAD_LINK_SECRET", "") or BOT_TOKEN or "dev-upload-link-secret"
    return secret.encode("utf-8")


def _sign_upload(filename: str, expires: int) -> str:
    payload = f"{filename}:{expires}".encode("utf-8")
    return hmac.new(_upload_signing_secret(), payload, hashlib.sha256).hexdigest()


def _verify_upload_token(filename: str, expires: int, token: str) -> None:
    if expires < int(time.time()):
        raise HTTPException(status_code=403, detail="Upload link expired.")
    expected = _sign_upload(filename, expires)
    if not hmac.compare_digest(expected, token or ""):
        raise HTTPException(status_code=403, detail="Invalid upload link.")


def _build_upload_url(stored_name: str) -> str:
    expires = int(time.time()) + UPLOAD_LINK_TTL_SECONDS
    query = urlencode({"expires": str(expires), "token": _sign_upload(stored_name, expires)})
    proof_url = f"/uploads/{stored_name}?{query}"
    if WEB_APP_URL:
        proof_url = f"{WEB_APP_URL}{proof_url}"
    return proof_url


def _is_upload_file(value: Any) -> bool:
    return isinstance(value, (UploadFile, StarletteUploadFile))


def _truthy_form_value(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "on", "accepted"}


async def _store_upload_file(file: StarletteUploadFile, *, label: str, allow_pdf: bool) -> Tuple[str, str]:
    mime = (file.content_type or "").lower()
    if allow_pdf:
        if mime not in {"image/jpeg", "image/png", "application/pdf"}:
            raise HTTPException(status_code=400, detail=f"Only JPG, PNG, or PDF is accepted for {label}.")
    elif mime not in {"image/jpeg", "image/png"}:
        raise HTTPException(status_code=400, detail=f"Only JPG or PNG is accepted for {label}.")

    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail=f"Uploaded {label} is empty.")
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File is too large. Max allowed size is {UPLOAD_MAX_MB:.1f} MB.",
            )
    finally:
        await file.close()

    actual_mime = _detect_upload_mime(content)
    if allow_pdf:
        allowed_actual = {"image/jpeg", "image/png", "application/pdf"}
    else:
        allowed_actual = {"image/jpeg", "image/png"}
    if actual_mime not in allowed_actual:
        raise HTTPException(status_code=400, detail=f"Uploaded {label} is not a valid JPG, PNG, or PDF.")
    if actual_mime != mime:
        raise HTTPException(status_code=400, detail=f"Uploaded {label} type does not match the file content.")

    suffix = { "application/pdf": ".pdf", "image/png": ".png", "image/jpeg": ".jpg" }[actual_mime]
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    stored_path = Path(UPLOAD_DIR) / stored_name
    try:
        stored_path.write_bytes(content)
    except OSError as exc:
        raise HTTPException(status_code=507, detail=f"Upload storage error: {exc}") from exc
    return _build_upload_url(stored_name), "external"


def _detect_upload_mime(content: bytes) -> str:
    if content.startswith(b"%PDF-"):
        return "application/pdf"
    if content.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image/png"
    if content.startswith(b"\xff\xd8\xff"):
        return "image/jpeg"
    return ""


def _client_rate_key(request: Request, scope: str, tg_id: Optional[int] = None) -> Tuple[str, str]:
    if tg_id is not None:
        return scope, f"tg:{tg_id}"
    forwarded = request.headers.get("x-forwarded-for", "")
    host = forwarded.split(",", 1)[0].strip()
    if not host and request.client:
        host = request.client.host
    return scope, host or "unknown"


def _enforce_rate_limit(request: Request, scope: str, limit: int, tg_id: Optional[int] = None) -> None:
    now = time.time()
    cutoff = now - RATE_LIMIT_WINDOW_SECONDS
    key = _client_rate_key(request, scope, tg_id)
    bucket = [ts for ts in _RATE_LIMIT_BUCKETS.get(key, []) if ts >= cutoff]
    if len(bucket) >= limit:
        raise HTTPException(status_code=429, detail="Too many requests. Try again shortly.")
    bucket.append(now)
    _RATE_LIMIT_BUCKETS[key] = bucket


def _email_login_configured() -> bool:
    return EMAIL_LOGIN_DEV_MODE or bool(RESEND_API_KEY and RESEND_FROM_EMAIL)


def _normalize_email(value: str) -> str:
    email = (value or "").strip().lower()
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
        raise HTTPException(status_code=400, detail="Enter a valid email address.")
    return email


def _email_login_secret() -> bytes:
    secret = (
        os.getenv("EMAIL_LOGIN_SECRET", "")
        or os.getenv("UPLOAD_LINK_SECRET", "")
        or BOT_TOKEN
        or ADMIN_WEB_PASSWORD
    )
    if not secret:
        secret = "dev-email-login-secret"
    return secret.encode("utf-8")


def _email_code_hash(email: str, code: str) -> str:
    normalized_code = "".join(ch for ch in str(code or "") if ch.isdigit())
    payload = f"{email}:{normalized_code}".encode("utf-8")
    return "hmac-sha256:" + hmac.new(_email_login_secret(), payload, hashlib.sha256).hexdigest()


def _send_login_code(email: str, code: str) -> None:
    if EMAIL_LOGIN_DEV_MODE:
        return
    if not RESEND_API_KEY or not RESEND_FROM_EMAIL:
        raise HTTPException(status_code=503, detail="Email login is not configured yet.")

    body = json.dumps(
        {
            "from": RESEND_FROM_EMAIL,
            "to": [email],
            "subject": "Your Budapest Tunderi login code",
            "text": (
                f"Your Budapest Tunderi login code is {code}. "
                f"It expires in {max(1, EMAIL_LOGIN_TTL_SECONDS // 60)} minutes."
            ),
            "html": (
                "<p>Your Budapest Tunderi login code is "
                f"<strong>{code}</strong>.</p>"
                f"<p>It expires in {max(1, EMAIL_LOGIN_TTL_SECONDS // 60)} minutes.</p>"
            ),
        }
    ).encode("utf-8")
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=body,
        method="POST",
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "BudapestTunderiTicketBot/1.0",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            if response.status >= 400:
                raise HTTPException(status_code=502, detail="Could not send the login code.")
    except urllib.error.HTTPError as exc:
        raise HTTPException(status_code=502, detail="Could not send the login code.") from exc
    except urllib.error.URLError as exc:
        raise HTTPException(status_code=502, detail="Email service is temporarily unavailable.") from exc


def _send_email(to_email: str, subject: str, text: str, html: Optional[str] = None) -> None:
    """Best-effort transactional email via Resend.

    Mirrors _send_login_code: no-op when email delivery is not configured or when
    EMAIL_LOGIN_DEV_MODE is on, and never raises into the caller (all network errors
    are swallowed) so booking approval/rejection flows are unaffected by mail failures.
    """
    recipient = (to_email or "").strip()
    if not recipient:
        return
    if EMAIL_LOGIN_DEV_MODE:
        return
    if not RESEND_API_KEY or not RESEND_FROM_EMAIL:
        return

    payload = {
        "from": RESEND_FROM_EMAIL,
        "to": [recipient],
        "subject": subject,
        "text": text,
    }
    if html:
        payload["html"] = html
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        data=body,
        method="POST",
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json",
            "User-Agent": "BudapestTunderiTicketBot/1.0",
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            _ = response.status
    except Exception:
        # Best-effort only: notifications must never break the admin action.
        return


def _verified_email_code_row(email: str, code_value: str):
    code = "".join(ch for ch in (code_value or "") if ch.isdigit())
    if len(code) != 6:
        raise HTTPException(status_code=400, detail="Enter the 6-digit code.")

    row = db.get_email_login_code(email)
    if not row:
        raise HTTPException(status_code=400, detail="Code expired. Request a new one.")
    try:
        expires_at = datetime.fromisoformat(row["expires_at"])
    except ValueError as exc:
        db.delete_email_login_code(email)
        raise HTTPException(status_code=400, detail="Code expired. Request a new one.") from exc
    if expires_at < datetime.now(timezone.utc):
        db.delete_email_login_code(email)
        raise HTTPException(status_code=400, detail="Code expired. Request a new one.")
    if int(row["attempts"] or 0) >= EMAIL_CODE_ATTEMPT_LIMIT:
        db.delete_email_login_code(email)
        raise HTTPException(status_code=429, detail="Too many wrong attempts. Request a new code.")

    expected = str(row["code_hash"])
    actual = _email_code_hash(email, code)
    if not hmac.compare_digest(expected, actual):
        attempts = db.increment_email_login_attempts(email)
        if attempts >= EMAIL_CODE_ATTEMPT_LIMIT:
            db.delete_email_login_code(email)
            raise HTTPException(status_code=429, detail="Too many wrong attempts. Request a new code.")
        raise HTTPException(status_code=400, detail="Wrong code. Check your email and try again.")
    return row


def _split_google_name(payload: Dict[str, Any]) -> Tuple[str, str]:
    given = str(payload.get("given_name") or "").strip()
    family = str(payload.get("family_name") or "").strip()
    if given:
        return given, family or "-"
    full = str(payload.get("name") or "").strip()
    parts = [part for part in full.split() if part]
    if not parts:
        return "Guest", "-"
    if len(parts) == 1:
        return parts[0], "-"
    return parts[0], " ".join(parts[1:])


def _verify_google_credential(credential: str) -> Dict[str, Any]:
    if not GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=503, detail="Google sign-in is not configured yet.")
    token = (credential or "").strip()
    if not token:
        raise HTTPException(status_code=400, detail="Google sign-in token is missing.")

    url = "https://oauth2.googleapis.com/tokeninfo?" + urlencode({"id_token": token})
    req = urllib.request.Request(
        url,
        headers={"User-Agent": "BudapestTunderiTicketBot/1.0"},
        method="GET",
    )
    try:
        with urllib.request.urlopen(req, timeout=10) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as exc:
        raise HTTPException(status_code=401, detail="Google sign-in token is invalid.") from exc
    except (urllib.error.URLError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=502, detail="Google sign-in is temporarily unavailable.") from exc

    if str(payload.get("aud") or "") != GOOGLE_CLIENT_ID:
        raise HTTPException(status_code=401, detail="Google sign-in token is for another app.")
    if str(payload.get("iss") or "") not in {"accounts.google.com", "https://accounts.google.com"}:
        raise HTTPException(status_code=401, detail="Google sign-in token issuer is invalid.")
    try:
        exp = int(payload.get("exp") or "0")
    except ValueError as exc:
        raise HTTPException(status_code=401, detail="Google sign-in token expiry is invalid.") from exc
    if exp < int(time.time()):
        raise HTTPException(status_code=401, detail="Google sign-in token expired.")
    if str(payload.get("email_verified") or "").lower() != "true":
        raise HTTPException(status_code=401, detail="Google email is not verified.")
    email = _normalize_email(str(payload.get("email") or ""))
    payload["email"] = email
    return payload


def _safe_upload_path(filename: str) -> Path:
    clean_name = Path(unquote(filename or "")).name
    if not clean_name or clean_name != filename or clean_name in {".", ".."}:
        raise HTTPException(status_code=404, detail="Upload not found.")
    file_path = (Path(UPLOAD_DIR) / clean_name).resolve()
    upload_root = Path(UPLOAD_DIR).resolve()
    if upload_root not in file_path.parents:
        raise HTTPException(status_code=404, detail="Upload not found.")
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Upload not found.")
    return file_path


def _safe_event_media_path(filename: str) -> Path:
    """Path-traversal-safe resolution for public event banners in EVENT_MEDIA_DIR."""
    clean_name = Path(unquote(filename or "")).name
    if not clean_name or clean_name != filename or clean_name in {".", ".."}:
        raise HTTPException(status_code=404, detail="Event media not found.")
    file_path = (Path(EVENT_MEDIA_DIR) / clean_name).resolve()
    media_root = Path(EVENT_MEDIA_DIR).resolve()
    if media_root not in file_path.parents:
        raise HTTPException(status_code=404, detail="Event media not found.")
    if not file_path.is_file():
        raise HTTPException(status_code=404, detail="Event media not found.")
    return file_path


def _event_media_filename(photo_url: str) -> Optional[str]:
    """Extract the stored filename from a '/event-media/<name>' banner URL."""
    raw = (photo_url or "").strip()
    if not raw:
        return None
    parsed = urlparse(raw)
    path = parsed.path if parsed.scheme else raw
    if not path.startswith("/event-media/"):
        return None
    filename = unquote(Path(path).name)
    if not filename or filename in {".", ".."}:
        return None
    return filename


def _delete_event_media(photo_url: str) -> None:
    """Best-effort removal of a stored public banner file."""
    filename = _event_media_filename(photo_url)
    if not filename:
        return
    try:
        _safe_event_media_path(filename).unlink()
    except HTTPException:
        return
    except OSError:
        return


def _absolute_media_url(url: str) -> str:
    """Return an absolute URL for a stored '/event-media/...' path when possible."""
    value = (url or "").strip()
    if not value:
        return ""
    if value.startswith("/") and WEB_APP_URL:
        return f"{WEB_APP_URL}{value}"
    return value


def _pending_status(status: str) -> bool:
    normalized = (status or "").strip().lower()
    return normalized in {
        STATUS_PENDING,
        "pending",
        "pending_payment",
        "pending_review",
        "pending_payment_approval",
    }


def cleanup_upload_storage(now_ts: Optional[float] = None) -> Dict[str, int]:
    upload_root = Path(UPLOAD_DIR)
    upload_root.mkdir(parents=True, exist_ok=True)
    now = float(now_ts if now_ts is not None else time.time())
    retention_seconds = UPLOAD_RETENTION_DAYS * 24 * 60 * 60

    referenced_any = set()
    referenced_pending = set()
    for row in [*db.list_external_payment_files(), *db.list_external_repost_files()]:
        filename = _extract_upload_filename(row["payment_file_id"])
        if not filename:
            continue
        referenced_any.add(filename)
        if _pending_status(row["status"]):
            referenced_pending.add(filename)

    scanned = 0
    deleted = 0
    kept = 0
    for file_path in upload_root.iterdir():
        if not file_path.is_file():
            continue
        scanned += 1
        filename = file_path.name
        if filename in referenced_pending:
            kept += 1
            continue

        orphan = filename not in referenced_any
        old = False
        if retention_seconds > 0:
            try:
                old = (now - file_path.stat().st_mtime) > retention_seconds
            except OSError:
                old = False

        if not (orphan or old):
            kept += 1
            continue

        try:
            file_path.unlink()
            deleted += 1
        except OSError:
            kept += 1

    return {
        "scanned": scanned,
        "deleted": deleted,
        "kept": kept,
        "referenced": len(referenced_any),
    }


def _maybe_run_upload_cleanup(force: bool = False) -> None:
    global _LAST_UPLOAD_CLEANUP_TS
    now = time.time()
    if not force and (now - _LAST_UPLOAD_CLEANUP_TS) < UPLOAD_CLEANUP_INTERVAL_SECONDS:
        return
    _LAST_UPLOAD_CLEANUP_TS = now
    try:
        cleanup_upload_storage(now_ts=now)
    except Exception:
        # Never break user flow because of cleanup issues.
        return


@app.on_event("startup")
def startup_cleanup() -> None:
    _maybe_run_upload_cleanup(force=True)


def _event_payload(event) -> Dict[str, Any]:
    tier = db.active_tier(event)
    payment_options = []
    for idx in (1, 2, 3):
        title = (getattr(event, f"payment{idx}_title", "") or "").strip()
        url = (getattr(event, f"payment{idx}_url", "") or "").strip()
        if not url:
            continue
        payment_options.append(
            {
                "slot": idx,
                "title": title or f"Payment Option {idx}",
                "url": url,
            }
        )
    return {
        "id": event.id,
        "title": event.title,
        "event_datetime": event.event_datetime,
        "location": event.location,
        "caption": event.caption,
        "photo_file_id": event.photo_file_id,
        "photo_url": _absolute_media_url(getattr(event, "photo_url", "") or ""),
        "maps_url": (getattr(event, "maps_url", "") or "").strip(),
        "repost_discount_enabled": event.repost_discount_enabled,
        "repost_discount_amount": event.repost_discount_amount,
        "girls_group_offer_enabled": event.girls_group_offer_enabled,
        "boys_group_offer_enabled": event.boys_group_offer_enabled,
        "tier": tier,
        "payment_options": payment_options,
        "payment": {
            "payment1_title": (event.payment1_title or "").strip(),
            "payment1_url": (event.payment1_url or "").strip(),
            "payment2_title": (event.payment2_title or "").strip(),
            "payment2_url": (event.payment2_url or "").strip(),
            "payment3_title": (event.payment3_title or "").strip(),
            "payment3_url": (event.payment3_url or "").strip(),
        },
    }


def _tier_label(tier_key: str) -> str:
    labels = {
        "early": "Early Bird",
        "tier1": "Regular Tier-1",
        "tier2": "Regular Tier-2",
    }
    return labels.get(tier_key, tier_key)


def _ticket_token_from_value(value: str) -> str:
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    parsed = urlparse(cleaned)
    if parsed.scheme and parsed.netloc:
        path_parts = [part for part in parsed.path.split("/") if part]
        if len(path_parts) >= 2 and path_parts[-2] == "checkin":
            return unquote(path_parts[-1]).strip()
        query = dict(parse_qsl(parsed.query, keep_blank_values=False))
        return (query.get("token") or "").strip()
    if cleaned.startswith("/checkin/"):
        return unquote(cleaned.split("/checkin/", 1)[1].split("?", 1)[0]).strip()
    return cleaned


def _ticket_payload(row) -> Dict[str, Any]:
    checked_in_at = row["checked_in_at"] if row else None
    return {
        "attendee_id": row["attendee_id"],
        "full_name": row["full_name"],
        "gender": row["gender"],
        "reservation_code": row["reservation_code"],
        "reservation_status": row["reservation_status"],
        "event_id": row["event_id"],
        "event_title": row["event_title"],
        "event_datetime": row["event_datetime"],
        "buyer_name": row["buyer_name"],
        "buyer_surname": row["buyer_surname"],
        "checked_in": bool((checked_in_at or "").strip()),
        "checked_in_at": checked_in_at,
        "checked_in_by_admin_tg_id": row["checked_in_by_admin_tg_id"],
    }


def _ticket_qr_text(request: Request, token: str) -> str:
    base_url = WEB_APP_URL or str(request.base_url).rstrip("/")
    return f"{base_url}/checkin/{token}"


def _ticket_qr_signature(token: str) -> str:
    secret = os.getenv("TICKET_QR_SECRET", "") or BOT_TOKEN or ADMIN_WEB_PASSWORD or "dev-ticket-qr-secret"
    return hmac.new(secret.encode("utf-8"), (token or "").strip().encode("utf-8"), hashlib.sha256).hexdigest()


def _bot_api(method: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    if not BOT_TOKEN:
        return {"ok": False, "description": "BOT_TOKEN is missing"}
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/{method}"
    req = urllib.request.Request(
        url,
        data=json.dumps(payload).encode("utf-8"),
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=12) as resp:
            body = resp.read().decode("utf-8")
            return json.loads(body)
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError):
        return {"ok": False}


def _notify_admins_pending_from_miniapp(reservation) -> None:
    event = db.get_event(reservation.event_id)
    user = db.get_user_by_id(reservation.user_id)
    attendees = db.list_attendees(reservation.id)
    attendee_lines = "\n".join([f"- {row['full_name']}" for row in attendees]) if attendees else "-"
    repost_lines = "\n".join(
        [
            f"- {row['full_name']}: {row['repost_proof_file_id']}"
            for row in attendees
            if int(row["repost_discount_applied"] or 0) == 1 and (row["repost_proof_file_id"] or "").strip()
        ]
    ) or "-"
    event_title = event.title if event else f"Event #{reservation.event_id}"
    tier_label = _tier_label(reservation.ticket_type)
    buyer = "Unknown"
    if user:
        buyer = f"{user.name} {user.surname} (tg:{user.tg_id})"
    applied_discount_amount = max(float(reservation.group_discount_amount or 0.0), float(reservation.discount_amount or 0.0))

    caption = (
        "New payment proof pending review\n\n"
        f"Code: {reservation.code}\n"
        f"Event: {event_title}\n"
        f"Tier: {tier_label}\n"
        f"Boys: {reservation.boys} | Girls: {reservation.girls}\n"
        f"Base total: {reservation.base_total_price:.2f}\n"
        f"Girls 2+1 discount: {reservation.girls_group_free_count} free = {reservation.girls_group_discount_amount:.2f}\n"
        f"Boys 3+1 discount: {reservation.boys_group_free_count} free = {reservation.boys_group_discount_amount:.2f}\n"
        f"Group offer discount total: {reservation.group_discount_amount:.2f}\n"
        f"Repost discount: {reservation.discount_count} x {reservation.discount_unit_amount:.2f} = {reservation.discount_amount:.2f}\n"
        f"Applied discount: {applied_discount_amount:.2f}\n"
        f"Final total: {reservation.total_price:.2f}\n"
        f"Buyer: {buyer}\n\n"
        f"Attendees:\n{attendee_lines}\n\n"
        f"Payment proof: {reservation.payment_file_id}\n\n"
        f"Repost proofs:\n{repost_lines}"
    )
    buttons = {
        "inline_keyboard": [
            [
                {"text": "Approve", "callback_data": f"review:approve:{reservation.id}"},
                {"text": "Reject Unreadable", "callback_data": f"review:reject:tpl:unreadable:{reservation.id}"},
            ],
            [
                {"text": "Reject Amount", "callback_data": f"review:reject:tpl:amount:{reservation.id}"},
                {"text": "Reject Custom", "callback_data": f"review:reject:custom:{reservation.id}"},
            ],
        ]
    }
    for admin_id in ADMIN_IDS:
        _bot_api(
            "sendMessage",
            {
                "chat_id": admin_id,
                "text": caption,
                "reply_markup": buttons,
                "disable_web_page_preview": False,
            },
        )


def _notify_user_pending_from_miniapp(reservation, tg_id: int) -> None:
    _bot_api(
        "sendMessage",
        {
            "chat_id": tg_id,
            "text": (
                "Your booking is pending admin approval.\n"
                f"Code: {reservation.code}\n\n"
                "Please wait while we verify your payment proof.\n"
                "Approval usually takes from 5 minutes up to 6 hours.\n"
                "If it takes longer, contact us directly on Telegram: @budapest_tunderi"
            ),
        },
    )


class QuoteRequest(BaseModel):
    event_id: int
    boys: int = Field(ge=0)
    girls: int = Field(ge=0)


class CheckInRequest(BaseModel):
    tg_id: Optional[int] = None
    token: str


class WebRegisterRequest(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    surname: str = Field(min_length=1, max_length=80)
    phone: str = Field(min_length=5, max_length=40)


class WebEmailLoginStartRequest(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    surname: str = Field(min_length=1, max_length=80)
    email: str = Field(min_length=5, max_length=254)
    phone: str = Field(min_length=5, max_length=40)


class WebEmailLoginVerifyRequest(BaseModel):
    email: str = Field(min_length=5, max_length=254)
    code: str = Field(min_length=4, max_length=12)


class WebEmailUpdateStartRequest(BaseModel):
    email: str = Field(min_length=5, max_length=254)


class WebEmailUpdateVerifyRequest(BaseModel):
    email: str = Field(min_length=5, max_length=254)
    code: str = Field(min_length=4, max_length=12)


class WebGoogleLoginRequest(BaseModel):
    credential: str = Field(min_length=20, max_length=4096)
    phone: str = Field(default="", max_length=40)


class WebProfileUpdateRequest(BaseModel):
    name: str = Field(min_length=1, max_length=80)
    surname: str = Field(min_length=1, max_length=80)
    phone: str = Field(min_length=5, max_length=40)


class WebCancelRequest(BaseModel):
    code: str = Field(min_length=1, max_length=64)


class AdminWebLoginRequest(BaseModel):
    password: str = Field(min_length=1, max_length=200)


class AdminGuestAddRequest(BaseModel):
    tg_id: Optional[int] = None
    reservation_code: str
    gender: str
    full_name: str


class AdminGuestRemoveRequest(BaseModel):
    tg_id: Optional[int] = None
    attendee_id: int


class AdminGuestRenameRequest(BaseModel):
    tg_id: Optional[int] = None
    attendee_id: int
    full_name: str


class AdminEventUpdateRequest(BaseModel):
    tg_id: Optional[int] = None
    event_id: int
    updates: Dict[str, Any]


class AdminEventDeleteRequest(BaseModel):
    tg_id: Optional[int] = None
    event_id: int


class AdminEventCreateSimpleRequest(BaseModel):
    tg_id: Optional[int] = None
    title: str
    caption: str = ""
    early_boy: float = Field(ge=0)
    early_girl: float = Field(ge=0)
    early_qty: int = Field(ge=0)
    tier1_boy: float = Field(ge=0)
    tier1_girl: float = Field(ge=0)
    tier1_qty: int = Field(ge=0)
    tier2_boy: float = Field(ge=0)
    tier2_girl: float = Field(ge=0)
    tier2_qty: int = Field(ge=0)
    repost_discount_enabled: bool = False
    repost_discount_amount: float = Field(default=0, ge=0)
    girls_group_offer_enabled: bool = False
    boys_group_offer_enabled: bool = False
    payment1_title: str = ""
    payment1_url: str = ""
    payment2_title: str = ""
    payment2_url: str = ""
    payment3_title: str = ""
    payment3_url: str = ""
    maps_url: str = ""
    location: Optional[str] = None
    event_datetime: Optional[str] = None


class AdminGuestAddByEventRequest(BaseModel):
    tg_id: Optional[int] = None
    event_id: int
    name: str
    surname: str
    gender: str


class AdminGuestRemoveByNameRequest(BaseModel):
    tg_id: Optional[int] = None
    event_id: int
    name: str
    surname: str


class AdminReservationApproveRequest(BaseModel):
    tg_id: Optional[int] = None
    reservation_id: int


class AdminReservationRejectRequest(BaseModel):
    tg_id: Optional[int] = None
    reservation_id: int
    note: str = ""


def _require_admin(tg_id: Optional[int]) -> int:
    if tg_id is None:
        raise HTTPException(status_code=401, detail="Missing tg_id.")
    if tg_id not in ADMIN_IDS:
        raise HTTPException(status_code=403, detail="Admin access denied.")
    return tg_id


def _request_admin(request: Request, provided_tg_id: Optional[int] = None) -> int:
    token = _request_admin_token(request)
    if token and db.is_valid_admin_web_session(token):
        return 0

    try:
        verified_tg_id = _request_tg_id(request, provided_tg_id)
    except HTTPException as exc:
        if exc.status_code == 401:
            raise HTTPException(status_code=401, detail="Admin login required.") from exc
        raise
    return _require_admin(verified_tg_id)


def _row_dict(row) -> Dict[str, Any]:
    return dict(row) if row is not None else {}


def _normalize_header_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().lower().replace("\ufeff", "")


def _parse_guest_row(row: tuple, row_index: int) -> Dict[str, Any]:
    value_name = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
    value_surname = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

    if not value_name and not value_surname:
        return {"skip": True, "reason": "empty"}

    if row_index == 1:
        h1 = _normalize_header_cell(value_name)
        h2 = _normalize_header_cell(value_surname)
        if h1 in {"name", "first_name", "firstname", "first name", "isim", "имя"} and h2 in {
            "surname",
            "last_name",
            "lastname",
            "last name",
            "soyad",
            "фамилия",
        }:
            return {"skip": True, "reason": "header"}

    if value_name and not value_surname and " " in value_name:
        parts = [p for p in value_name.split() if p]
        if len(parts) >= 2:
            value_name = parts[0]
            value_surname = " ".join(parts[1:])

    if not value_name and value_surname:
        value_name = value_surname
        value_surname = ""

    if not value_name:
        return {"skip": True, "reason": "missing_name"}

    return {
        "skip": False,
        "name": value_name,
        "surname": value_surname,
    }


@app.get("/")
def root() -> FileResponse:
    return FileResponse(WEB_DIR / "index.html")


@app.get("/admin")
def admin_page() -> RedirectResponse:
    return RedirectResponse(url="/?open_admin=1", status_code=307)


@app.get("/checkin/{token}")
def checkin_page(token: str) -> RedirectResponse:
    encoded = urlencode({"open_admin": "1", "checkin": token})
    return RedirectResponse(url=f"/?{encoded}", status_code=307)


@app.get("/health")
def health() -> Dict[str, str]:
    return {"status": "ok"}


@app.get("/robots.txt")
def robots_txt() -> PlainTextResponse:
    sitemap_target = f"{WEB_APP_URL}/sitemap.xml" if WEB_APP_URL else "/sitemap.xml"
    lines = [
        "User-agent: *",
        "Allow: /",
        "Disallow: /api/",
        "Disallow: /admin",
        "Disallow: /uploads/",
        "Disallow: /checkin/",
        f"Sitemap: {sitemap_target}",
    ]
    return PlainTextResponse("\n".join(lines) + "\n")


@app.get("/sitemap.xml")
def sitemap_xml() -> Response:
    home_url = WEB_APP_URL + "/" if WEB_APP_URL else "/"
    xml = (
        '<?xml version="1.0" encoding="UTF-8"?>\n'
        '<urlset xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f"  <url>\n    <loc>{home_url}</loc>\n    <changefreq>weekly</changefreq>\n"
        "    <priority>1.0</priority>\n  </url>\n"
        "</urlset>\n"
    )
    return Response(content=xml, media_type="application/xml")


_NOT_FOUND_HTML = """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>Page not found — Budapest Tunderi</title>
  <style>
    body {
      margin: 0;
      min-height: 100vh;
      display: flex;
      flex-direction: column;
      align-items: center;
      justify-content: center;
      gap: 12px;
      font-family: "Manrope", "Segoe UI", Arial, sans-serif;
      color: #f6f0e6;
      background: linear-gradient(180deg, #08080f 0%, #151019 48%, #0f0c12 100%);
      text-align: center;
      padding: 24px;
    }
    h1 { margin: 0; font-size: 3rem; color: #d8a83f; }
    p { margin: 0; color: #b8aa99; }
    a { color: #d8a83f; text-decoration: none; }
    a:hover { text-decoration: underline; }
  </style>
</head>
<body>
  <h1>404</h1>
  <p>This page could not be found.</p>
  <p><a href="/">Return to Budapest Tunderi</a></p>
</body>
</html>
"""


@app.exception_handler(StarletteHTTPException)
async def not_found_handler(request: Request, exc: StarletteHTTPException):
    if exc.status_code == 404 and not request.url.path.startswith("/api/"):
        return HTMLResponse(content=_NOT_FOUND_HTML, status_code=404)
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail},
        headers=getattr(exc, "headers", None),
    )


@app.get("/uploads/{filename}")
def signed_upload(filename: str, expires: int, token: str) -> FileResponse:
    _verify_upload_token(filename, expires, token)
    file_path = _safe_upload_path(filename)
    suffix = file_path.suffix.lower()
    media_type = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
        ".pdf": "application/pdf",
    }.get(suffix, "application/octet-stream")
    return FileResponse(
        file_path,
        media_type=media_type,
        headers={
            "Cache-Control": "private, max-age=300",
            "Content-Disposition": f'inline; filename="{file_path.name}"',
        },
    )


@app.get("/event-media/{filename}")
def event_media(filename: str) -> FileResponse:
    # Public + permanent event banners. Unsigned by design (unlike payment proofs).
    file_path = _safe_event_media_path(filename)
    suffix = file_path.suffix.lower()
    media_type = {
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png": "image/png",
    }.get(suffix, "application/octet-stream")
    return FileResponse(
        file_path,
        media_type=media_type,
        headers={
            "Cache-Control": "public, max-age=31536000, immutable",
            "Content-Disposition": f'inline; filename="{file_path.name}"',
        },
    )


@app.get("/api/events")
def list_events() -> Dict[str, Any]:
    items = []
    for event in db.list_events():
        payload = _event_payload(event)
        if payload["tier"]:
            items.append(payload)
    return {"items": items}


def _profile_payload(user) -> Dict[str, Any]:
    return {
        "tg_id": user.tg_id,
        "name": user.name,
        "surname": user.surname,
        "email": getattr(user, "email", "") or "",
        "phone": user.phone,
        "source": "telegram" if int(user.tg_id) > 0 else "website",
    }


@app.get("/api/web/auth_config")
def web_auth_config() -> Dict[str, Any]:
    return {
        "email_login_enabled": _email_login_configured(),
        "code_ttl_seconds": EMAIL_LOGIN_TTL_SECONDS,
        "google_client_id": GOOGLE_CLIENT_ID,
    }


@app.post("/api/web/register")
def web_register(request: Request, response: Response, payload: WebRegisterRequest) -> Dict[str, Any]:
    if _email_login_configured() and not LEGACY_WEB_REGISTER_ENABLED:
        raise HTTPException(status_code=410, detail="Use email verification to continue.")
    _enforce_rate_limit(request, "web_register", 20)
    try:
        existing_user = db.get_user_by_web_session(_request_web_token(request))
        if existing_user:
            _ensure_not_blocked(existing_user)
            user = db.update_web_user_profile(
                existing_user.id,
                payload.name.strip(),
                payload.surname.strip(),
                payload.phone.strip(),
            )
            token = _request_web_token(request)
        else:
            user, token = db.create_web_user(
                payload.name.strip(),
                payload.surname.strip(),
                payload.phone.strip(),
            )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    _set_session_cookie(request, response, WEB_SESSION_COOKIE, token)
    return {"ok": True, "profile": _profile_payload(user)}


@app.post("/api/web/login/start")
def web_email_login_start(request: Request, payload: WebEmailLoginStartRequest) -> Dict[str, Any]:
    if not _email_login_configured():
        raise HTTPException(status_code=503, detail="Email login is not configured yet.")
    _enforce_rate_limit(request, "web_email_login", EMAIL_LOGIN_RATE_LIMIT)
    email = _normalize_email(payload.email)
    code = f"{secrets.randbelow(1_000_000):06d}"
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=EMAIL_LOGIN_TTL_SECONDS)
    db.save_email_login_code(
        email=email,
        code_hash=_email_code_hash(email, code),
        name=payload.name.strip(),
        surname=payload.surname.strip(),
        phone=payload.phone.strip(),
        expires_at=expires_at.isoformat(),
    )
    _send_login_code(email, code)
    response: Dict[str, Any] = {"ok": True, "message": "Code sent."}
    if EMAIL_LOGIN_DEV_MODE:
        response["dev_code"] = code
    return response


@app.post("/api/web/login/verify")
def web_email_login_verify(request: Request, response: Response, payload: WebEmailLoginVerifyRequest) -> Dict[str, Any]:
    if not _email_login_configured():
        raise HTTPException(status_code=503, detail="Email login is not configured yet.")
    _enforce_rate_limit(request, "web_email_verify", EMAIL_LOGIN_RATE_LIMIT)
    email = _normalize_email(payload.email)
    row = _verified_email_code_row(email, payload.code)

    try:
        user, token = db.create_or_update_web_user_by_email(
            row["name"],
            row["surname"],
            email,
            row["phone"],
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        db.delete_email_login_code(email)
    _set_session_cookie(request, response, WEB_SESSION_COOKIE, token)
    return {"ok": True, "profile": _profile_payload(user)}


@app.post("/api/web/email/start")
def web_email_update_start(request: Request, payload: WebEmailUpdateStartRequest) -> Dict[str, Any]:
    if not _email_login_configured():
        raise HTTPException(status_code=503, detail="Email login is not configured yet.")
    user, _verified_tg_id = _request_user(request, None)
    _enforce_rate_limit(request, "web_email_update", EMAIL_LOGIN_RATE_LIMIT, tg_id=user.tg_id)
    email = _normalize_email(payload.email)
    existing = db.get_user_by_email(email)
    if existing and int(existing.id) != int(user.id):
        raise HTTPException(status_code=409, detail="This email is already used by another account.")
    code = f"{secrets.randbelow(1_000_000):06d}"
    expires_at = datetime.now(timezone.utc) + timedelta(seconds=EMAIL_LOGIN_TTL_SECONDS)
    db.save_email_login_code(
        email=email,
        code_hash=_email_code_hash(email, code),
        name=user.name,
        surname=user.surname,
        phone=user.phone,
        expires_at=expires_at.isoformat(),
    )
    _send_login_code(email, code)
    response: Dict[str, Any] = {"ok": True, "message": "Code sent."}
    if EMAIL_LOGIN_DEV_MODE:
        response["dev_code"] = code
    return response


@app.post("/api/web/email/verify")
def web_email_update_verify(
    request: Request,
    response: Response,
    payload: WebEmailUpdateVerifyRequest,
) -> Dict[str, Any]:
    if not _email_login_configured():
        raise HTTPException(status_code=503, detail="Email login is not configured yet.")
    user, _verified_tg_id = _request_user(request, None)
    _enforce_rate_limit(request, "web_email_update_verify", EMAIL_LOGIN_RATE_LIMIT, tg_id=user.tg_id)
    email = _normalize_email(payload.email)
    row = _verified_email_code_row(email, payload.code)
    try:
        updated = db.update_web_user_email(user.id, email)
    except ValueError as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    finally:
        db.delete_email_login_code(row["email"])
    token = _request_web_token(request)
    if token:
        _set_session_cookie(request, response, WEB_SESSION_COOKIE, token)
    return {"ok": True, "profile": _profile_payload(updated)}


@app.post("/api/web/login/google")
def web_google_login(request: Request, response: Response, payload: WebGoogleLoginRequest) -> Dict[str, Any]:
    _enforce_rate_limit(request, "web_google_login", EMAIL_LOGIN_RATE_LIMIT)
    google_payload = _verify_google_credential(payload.credential)
    name, surname = _split_google_name(google_payload)
    try:
        user, token = db.create_or_update_web_user_by_email(
            name,
            surname,
            google_payload["email"],
            payload.phone.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    _set_session_cookie(request, response, WEB_SESSION_COOKIE, token)
    return {"ok": True, "profile": _profile_payload(user)}


@app.put("/api/web/profile")
def web_profile_update(request: Request, response: Response, payload: WebProfileUpdateRequest) -> Dict[str, Any]:
    user, _verified_tg_id = _request_user(request, None)
    try:
        updated = db.update_web_user_profile(
            user.id,
            payload.name.strip(),
            payload.surname.strip(),
            payload.phone.strip(),
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    token = _request_web_token(request)
    if token:
        _set_session_cookie(request, response, WEB_SESSION_COOKIE, token)
    return {"ok": True, "profile": _profile_payload(updated)}


@app.post("/api/web/logout")
def web_logout(request: Request, response: Response) -> Dict[str, Any]:
    token = _request_web_token(request)
    if token:
        db.delete_web_session(token)
    response.delete_cookie(WEB_SESSION_COOKIE, path="/")
    return {"ok": True}


@app.post("/api/web/cancel")
def web_cancel(
    request: Request,
    payload: WebCancelRequest,
    tg_id: Optional[int] = None,
) -> Dict[str, Any]:
    user, _verified_tg_id = _request_user(request, tg_id)
    code = (payload.code or "").strip()
    reservation = db.get_reservation_by_code(code)
    if not reservation or int(reservation.user_id) != int(user.id):
        raise HTTPException(status_code=404, detail="Reservation not found for your account.")
    if (reservation.status or "").strip() != STATUS_PENDING:
        raise HTTPException(
            status_code=400,
            detail="Approved bookings can't be cancelled online. Please contact us.",
        )
    ok, message, updated = db.cancel_reservation_for_user(user.id, code)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "status": updated.status if updated else None}


@app.get("/api/me")
def me(request: Request, tg_id: Optional[int] = None) -> Dict[str, Any]:
    user, _verified_tg_id = _request_user(request, tg_id)
    return {"profile": _profile_payload(user)}


@app.get("/api/my_tickets")
def my_tickets(request: Request, tg_id: Optional[int] = None, limit: int = 20) -> Dict[str, Any]:
    user, _verified_tg_id = _request_user(request, tg_id)
    rows = db.list_reservations_for_user(user.id)[: max(1, min(limit, 100))]
    items = []
    for reservation in rows:
        event = db.get_event(reservation.event_id)
        attendees = db.list_attendees(reservation.id)
        ticket_items = []
        for row in attendees:
            checked_in_at = row["checked_in_at"]
            ticket_item = {
                "attendee_id": row["id"],
                "full_name": row["full_name"],
                "status": row["status"],
                "checked_in": bool((checked_in_at or "").strip()),
                "checked_in_at": checked_in_at,
            }
            if (reservation.status or "").strip().lower() == "approved":
                token = row["ticket_token"]
                ticket_item["qr_url"] = f"/api/tickets/{token}/qr?sig={_ticket_qr_signature(token)}"
            ticket_items.append(ticket_item)
        items.append(
            {
                "code": reservation.code,
                "event_id": reservation.event_id,
                "event_title": event.title if event else f"Event #{reservation.event_id}",
                "status": reservation.status,
                "admin_note": reservation.admin_note or "",
                "tier_label": _tier_label(reservation.ticket_type),
                "boys": reservation.boys,
                "girls": reservation.girls,
                "total_price": reservation.total_price,
                "attendees": [row["full_name"] for row in attendees],
                "tickets": ticket_items,
            }
        )
    return {"items": items}


@app.get("/api/tickets/{token}/qr")
def ticket_qr(request: Request, token: str, sig: Optional[str] = None, tg_id: Optional[int] = None) -> Response:
    clean_token = _ticket_token_from_value(token)
    row = db.lookup_ticket(clean_token)
    if not row:
        raise HTTPException(status_code=404, detail="Ticket not found.")

    allowed = bool(sig and hmac.compare_digest(sig, _ticket_qr_signature(clean_token)))
    if not allowed:
        try:
            user, _verified_tg_id = _request_user(request, tg_id)
            allowed = int(row["user_id"]) == int(user.id)
        except HTTPException:
            try:
                _request_admin(request, tg_id)
                allowed = True
            except HTTPException:
                allowed = False
    if not allowed:
        raise HTTPException(status_code=403, detail="Ticket access denied.")
    if (row["reservation_status"] or "").strip().lower() != "approved":
        raise HTTPException(status_code=403, detail="Ticket is not approved yet.")

    image = qrcode.make(_ticket_qr_text(request, row["ticket_token"]))
    out = BytesIO()
    image.save(out, format="PNG")
    return Response(
        content=out.getvalue(),
        media_type="image/png",
        headers={"Cache-Control": "private, no-store, max-age=0"},
    )


@app.post("/api/book_with_payment")
async def book_with_payment(request: Request) -> Dict[str, Any]:
    _maybe_run_upload_cleanup()
    form = await request.form()
    upload_values = [value for _, value in form.multi_items() if _is_upload_file(value)]
    try:
        try:
            event_id = int(str(form.get("event_id", "")).strip())
            boys = int(str(form.get("boys", "")).strip())
            girls = int(str(form.get("girls", "")).strip())
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="event_id, boys, and girls must be integers.") from exc
        if boys < 0 or girls < 0:
            raise HTTPException(status_code=400, detail="Boys and girls must be non-negative.")
        raw_tg_id = str(form.get("tg_id", "")).strip()
        try:
            provided_tg_id = int(raw_tg_id) if raw_tg_id else None
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="tg_id must be an integer when provided.") from exc
        user, verified_tg_id = _request_user(request, provided_tg_id)
        if int(getattr(user, "tg_id", 0) or 0) <= 0 and not (getattr(user, "phone", "") or "").strip():
            raise HTTPException(
                status_code=400,
                detail="Add your phone number to your profile before booking.",
            )
        _enforce_rate_limit(request, "booking", BOOKING_RATE_LIMIT, verified_tg_id or user.id)
        if not _truthy_form_value(form.get("terms_accepted")):
            raise HTTPException(status_code=400, detail="Accept the booking terms before booking.")
        attendees = str(form.get("attendees", ""))
        payment_file = form.get("file")
        if not _is_upload_file(payment_file):
            raise HTTPException(status_code=400, detail="Payment proof file is required.")

        try:
            attendees_list = json.loads(attendees)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="attendees must be valid JSON list.") from exc
        if not isinstance(attendees_list, list):
            raise HTTPException(status_code=400, detail="attendees must be a list.")
        normalized_attendees = [str(x).strip() for x in attendees_list]
        if any(len(name.split()) < 2 for name in normalized_attendees):
            raise HTTPException(status_code=400, detail='Each attendee must be in format "Name Surname".')

        event = db.get_event(event_id)
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")

        discounted_indexes_raw = str(form.get("discounted_attendee_indexes", "[]"))
        try:
            discounted_indexes_payload = json.loads(discounted_indexes_raw)
        except json.JSONDecodeError as exc:
            raise HTTPException(status_code=400, detail="discounted_attendee_indexes must be valid JSON list.") from exc
        if not isinstance(discounted_indexes_payload, list):
            raise HTTPException(status_code=400, detail="discounted_attendee_indexes must be a list.")
        try:
            discounted_indexes = sorted({int(idx) for idx in discounted_indexes_payload})
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail="discounted_attendee_indexes must contain integers.") from exc
        if any(idx < 0 or idx >= len(normalized_attendees) for idx in discounted_indexes):
            raise HTTPException(status_code=400, detail="Discounted attendee indexes are out of range.")
        if discounted_indexes and not bool(event.repost_discount_enabled):
            raise HTTPException(status_code=400, detail="Repost discount is not enabled for this event.")
        try:
            db.quote_booking(event_id, boys, girls)
        except ValueError as exc:
            text = str(exc)
            if text == "Event not found":
                raise HTTPException(status_code=404, detail=text) from exc
            if "sold out" in text.lower() or "not enough tickets" in text.lower():
                raise HTTPException(status_code=409, detail=text) from exc
            raise HTTPException(status_code=400, detail=text) from exc

        repost_proofs_by_index: Dict[int, Tuple[str, str]] = {}
        stored_upload_urls: List[str] = []
        for idx in discounted_indexes:
            repost_file = form.get(f"repost_file_{idx}")
            if not _is_upload_file(repost_file):
                raise HTTPException(status_code=400, detail="Upload repost screenshot for each selected attendee.")
            repost_proofs_by_index[idx] = await _store_upload_file(
                repost_file,
                label=f"repost screenshot for attendee #{idx + 1}",
                allow_pdf=False,
            )
            stored_upload_urls.append(repost_proofs_by_index[idx][0])

        proof_url, proof_type = await _store_upload_file(
            payment_file,
            label="payment proof",
            allow_pdf=True,
        )
        stored_upload_urls.append(proof_url)

        try:
            reservation = db.create_pending_reservation(
                user_id=user.id,
                event_id=int(event_id),
                boys=int(boys),
                girls=int(girls),
                attendees=normalized_attendees,
                payment_file_id=proof_url,
                payment_file_type=proof_type,
                discounted_attendee_indexes=discounted_indexes,
                repost_proofs_by_index=repost_proofs_by_index,
            )
        except ValueError as exc:
            for upload_url in stored_upload_urls:
                _delete_stored_upload(upload_url)
            raise HTTPException(status_code=400, detail=str(exc)) from exc

        _notify_admins_pending_from_miniapp(reservation)
        if verified_tg_id is not None and verified_tg_id > 0:
            _notify_user_pending_from_miniapp(reservation, verified_tg_id)
        return {
            "ok": True,
            "code": reservation.code,
            "status": reservation.status,
        }
    finally:
        for upload in upload_values:
            try:
                await upload.close()
            except Exception:
                continue


@app.post("/api/quote")
def quote(request: Request, payload: QuoteRequest) -> Dict[str, Any]:
    _enforce_rate_limit(request, "quote", QUOTE_RATE_LIMIT)
    try:
        return db.quote_booking(payload.event_id, payload.boys, payload.girls)
    except ValueError as exc:
        text = str(exc)
        if text == "Event not found":
            raise HTTPException(status_code=404, detail=text) from exc
        if "sold out" in text.lower() or "not enough tickets" in text.lower():
            raise HTTPException(status_code=409, detail=text) from exc
        raise HTTPException(status_code=400, detail=text) from exc


@app.post("/api/admin/login")
def admin_login(request: Request, response: Response, payload: AdminWebLoginRequest) -> Dict[str, Any]:
    _enforce_rate_limit(request, "admin_login", 10)
    if not ADMIN_WEB_PASSWORD:
        raise HTTPException(status_code=503, detail="Website admin login is not configured.")
    if not hmac.compare_digest(payload.password, ADMIN_WEB_PASSWORD):
        raise HTTPException(status_code=403, detail="Wrong admin password.")
    token = db.create_admin_web_session()
    _set_session_cookie(request, response, ADMIN_SESSION_COOKIE, token)
    return {"ok": True}


@app.get("/api/admin/bootstrap")
def admin_bootstrap(request: Request, tg_id: Optional[int] = None) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, tg_id)
    return {
        "ok": True,
        "tg_id": verified_tg_id if verified_tg_id else None,
        "source": "telegram" if verified_tg_id else "website",
    }


@app.get("/api/admin/guests")
def admin_guests(
    request: Request,
    tg_id: Optional[int] = None,
    sort_by: str = "newest",
    search: Optional[str] = None,
    limit: Optional[int] = None,
) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, tg_id)
    rows = db.list_guests(sort_by=sort_by, search=search, limit=limit)
    return {"items": [_row_dict(r) for r in rows]}


@app.get("/api/admin/reservations")
def admin_reservations(
    request: Request,
    tg_id: Optional[int] = None,
    search: Optional[str] = None,
    limit: int = 25,
) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, tg_id)
    rows = db.list_active_reservations(search=search, limit=limit)
    return {"items": [_row_dict(r) for r in rows]}


@app.get("/api/admin/checkin/lookup")
def admin_checkin_lookup(request: Request, token: str, tg_id: Optional[int] = None) -> Dict[str, Any]:
    _request_admin(request, tg_id)
    clean_token = _ticket_token_from_value(token)
    row = db.lookup_ticket(clean_token)
    if not row:
        raise HTTPException(status_code=404, detail="Ticket not found.")
    return {"ok": True, "ticket": _ticket_payload(row)}


@app.post("/api/admin/checkin")
def admin_checkin(request: Request, payload: CheckInRequest) -> Dict[str, Any]:
    admin_tg_id = _request_admin(request, payload.tg_id)
    clean_token = _ticket_token_from_value(payload.token)
    ok, message, row = db.check_in_ticket(clean_token, admin_tg_id)
    if not row:
        raise HTTPException(status_code=404, detail=message)
    if not ok and message == "Ticket is not approved yet.":
        raise HTTPException(status_code=409, detail=message)
    if not ok and message == "Ticket already checked in.":
        return {"ok": False, "message": message, "ticket": _ticket_payload(row)}
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "ticket": _ticket_payload(row)}


@app.get("/api/admin/events")
def admin_events(request: Request, tg_id: Optional[int] = None) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, tg_id)
    items = []
    for event in db.list_events():
        payload = _event_payload(event)
        payload["prices"] = {
            "early_boy": event.early_bird_price,
            "early_girl": event.early_bird_price_girl,
            "early_qty": event.early_bird_qty,
            "tier1_boy": event.regular_tier1_price,
            "tier1_girl": event.regular_tier1_price_girl,
            "tier1_qty": event.regular_tier1_qty,
            "tier2_boy": event.regular_tier2_price,
            "tier2_girl": event.regular_tier2_price_girl,
            "tier2_qty": event.regular_tier2_qty,
            "repost_discount_enabled": event.repost_discount_enabled,
            "repost_discount_amount": event.repost_discount_amount,
            "girls_group_offer_enabled": event.girls_group_offer_enabled,
            "boys_group_offer_enabled": event.boys_group_offer_enabled,
        }
        items.append(payload)
    return {"items": items}


@app.post("/api/admin/guest/add")
def admin_guest_add(request: Request, payload: AdminGuestAddRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message, reservation = db.admin_add_guest(
        reservation_code=payload.reservation_code.strip(),
        full_name=payload.full_name.strip(),
        gender_raw=payload.gender.strip().lower(),
    )
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/guest/remove")
def admin_guest_remove(request: Request, payload: AdminGuestRemoveRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message, reservation = db.admin_remove_guest(payload.attendee_id)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/guest/rename")
def admin_guest_rename(request: Request, payload: AdminGuestRenameRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message = db.admin_rename_guest(payload.attendee_id, payload.full_name.strip())
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message}


@app.post("/api/admin/guest/add_by_event")
def admin_guest_add_by_event(request: Request, payload: AdminGuestAddByEventRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message, reservation = db.admin_add_guest_by_event(
        admin_tg_id=verified_tg_id,
        event_id=payload.event_id,
        name=payload.name.strip(),
        surname=payload.surname.strip(),
        gender_raw=payload.gender.strip().lower(),
    )
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/guest/remove_by_name")
def admin_guest_remove_by_name(request: Request, payload: AdminGuestRemoveByNameRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message, reservation = db.admin_remove_guest_by_name(
        event_id=payload.event_id,
        name=payload.name.strip(),
        surname=payload.surname.strip(),
    )
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/guest/import_xlsx")
async def admin_guest_import_xlsx(
    request: Request,
    tg_id: Optional[int] = Form(None),
    event_id: int = Form(...),
    file: UploadFile = File(...),
) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, tg_id)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload .xlsx file.")

    try:
        raw = await file.read()
    finally:
        await file.close()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File is too large. Max allowed size is {UPLOAD_MAX_MB:.1f} MB.",
        )
    try:
        workbook = load_workbook(filename=BytesIO(raw), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Failed to parse xlsx: {exc}") from exc

    sheet = workbook.active
    added = 0
    skipped = 0
    errors = []

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        parsed = _parse_guest_row(row, row_index)
        if parsed["skip"]:
            if parsed["reason"] in {"empty", "header"}:
                continue
            skipped += 1
            errors.append(f"Row {row_index}: invalid name/surname values.")
            continue

        value_name = parsed["name"]
        value_surname = parsed["surname"]

        ok, message, _reservation = db.admin_import_guest_by_event(
            admin_tg_id=verified_tg_id,
            event_id=event_id,
            name=value_name,
            surname=value_surname,
        )
        if ok:
            added += 1
        else:
            skipped += 1
            errors.append(f"Row {row_index}: {message}")

    return {
        "ok": True,
        "added": added,
        "skipped": skipped,
        "errors": errors[:10],
    }


@app.get("/api/admin/guest/export_xlsx")
def admin_guest_export_xlsx(request: Request, tg_id: Optional[int] = None) -> StreamingResponse:
    verified_tg_id = _request_admin(request, tg_id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Guests"
    sheet.append(["Name", "Surname"])
    for first, last in db.list_guest_name_pairs():
        sheet.append([first, last])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="guests_export.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/api/admin/event/update")
def admin_event_update(request: Request, payload: AdminEventUpdateRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    ok, message = db.set_event_fields(event_id=payload.event_id, updates=payload.updates)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    event = db.get_event(payload.event_id)
    return {"ok": True, "message": message, "event": event.__dict__ if event else None}


@app.post("/api/admin/event/delete")
def admin_event_delete(request: Request, payload: AdminEventDeleteRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    existing = db.get_event(payload.event_id)
    photo_url = getattr(existing, "photo_url", "") if existing else ""
    ok, message, deleted = db.delete_event(event_id=payload.event_id)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    # Best-effort removal of the public banner file for the deleted event.
    if photo_url:
        _delete_event_media(photo_url)
    return {"ok": True, "message": message, "deleted": deleted}


@app.post("/api/admin/event/create_simple")
def admin_event_create_simple(request: Request, payload: AdminEventCreateSimpleRequest) -> Dict[str, Any]:
    verified_tg_id = _request_admin(request, payload.tg_id)
    title = payload.title.strip()
    if not title:
        raise HTTPException(status_code=400, detail="Title is required.")

    # Payment fields accept a URL OR free text (e.g. a phone number for
    # Revolut/bank transfer), so there is no scheme requirement here.
    # The maps field accepts an https URL OR a pasted Google Maps <iframe> embed
    # (we extract the src); non-https values are rejected.
    try:
        maps_url = normalize_maps_url(payload.maps_url)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    total_qty = int(payload.early_qty) + int(payload.tier1_qty) + int(payload.tier2_qty)
    if total_qty <= 0:
        raise HTTPException(status_code=400, detail="At least one ticket quantity must be greater than 0.")

    location = (payload.location or "Budapest").strip() or "Budapest"
    if payload.event_datetime:
        event_datetime = payload.event_datetime.strip()
        try:
            db.parse_event_datetime(event_datetime)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Invalid datetime format. Use YYYY-MM-DD HH:MM") from exc
    else:
        default_dt = (datetime.now(ZoneInfo("Europe/Budapest")) + timedelta(days=7)).replace(
            minute=0,
            second=0,
            microsecond=0,
        )
        event_datetime = default_dt.strftime("%Y-%m-%d %H:%M")

    event_id = db.create_event(
        title=title,
        event_datetime=event_datetime,
        location=location,
        caption=payload.caption.strip(),
        photo_file_id="",
        early_boy_price=float(payload.early_boy),
        early_girl_price=float(payload.early_girl),
        early_qty=int(payload.early_qty),
        tier1_boy_price=float(payload.tier1_boy),
        tier1_girl_price=float(payload.tier1_girl),
        tier1_qty=int(payload.tier1_qty),
        tier2_boy_price=float(payload.tier2_boy),
        tier2_girl_price=float(payload.tier2_girl),
        tier2_qty=int(payload.tier2_qty),
        repost_discount_enabled=bool(payload.repost_discount_enabled),
        repost_discount_amount=float(payload.repost_discount_amount),
        girls_group_offer_enabled=bool(payload.girls_group_offer_enabled),
        boys_group_offer_enabled=bool(payload.boys_group_offer_enabled),
        payment1_title=(payload.payment1_title or "").strip(),
        payment1_url=(payload.payment1_url or "").strip(),
        payment2_title=(payload.payment2_title or "").strip(),
        payment2_url=(payload.payment2_url or "").strip(),
        payment3_title=(payload.payment3_title or "").strip(),
        payment3_url=(payload.payment3_url or "").strip(),
        maps_url=maps_url,
    )
    event = db.get_event(event_id)
    return {
        "ok": True,
        "message": "Event created.",
        "event": event.__dict__ if event else None,
    }


@app.post("/api/admin/event/photo")
async def admin_event_photo(
    request: Request,
    tg_id: Optional[int] = Form(None),
    event_id: int = Form(...),
    file: UploadFile = File(...),
) -> Dict[str, Any]:
    _request_admin(request, tg_id)

    event = db.get_event(event_id)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found.")

    mime = (file.content_type or "").lower()
    if mime not in {"image/jpeg", "image/png"}:
        raise HTTPException(status_code=400, detail="Only JPG or PNG is accepted for the event banner.")
    try:
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Uploaded banner is empty.")
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(
                status_code=413,
                detail=f"File is too large. Max allowed size is {UPLOAD_MAX_MB:.1f} MB.",
            )
    finally:
        await file.close()

    actual_mime = _detect_upload_mime(content)
    if actual_mime not in {"image/jpeg", "image/png"}:
        raise HTTPException(status_code=400, detail="Uploaded banner is not a valid JPG or PNG.")
    if actual_mime != mime:
        raise HTTPException(status_code=400, detail="Uploaded banner type does not match the file content.")

    suffix = {"image/png": ".png", "image/jpeg": ".jpg"}[actual_mime]
    stored_name = f"{uuid.uuid4().hex}{suffix}"
    stored_path = Path(EVENT_MEDIA_DIR) / stored_name
    try:
        stored_path.write_bytes(content)
    except OSError as exc:
        raise HTTPException(status_code=507, detail=f"Banner storage error: {exc}") from exc

    # Remove the previous banner (if any) once the new one is safely stored.
    old_photo_url = getattr(event, "photo_url", "") or ""
    new_photo_url = f"/event-media/{stored_name}"
    ok, message = db.set_event_photo(event_id, new_photo_url)
    if not ok:
        _delete_event_media(new_photo_url)
        raise HTTPException(status_code=400, detail=message)
    if old_photo_url and old_photo_url != new_photo_url:
        _delete_event_media(old_photo_url)

    updated = db.get_event(event_id)
    return {"ok": True, "message": "Event banner updated.", "event": updated.__dict__ if updated else None}


def _pending_reservation_item(row) -> Dict[str, Any]:
    data = _row_dict(row)
    payment_file_id = data.get("payment_file_id") or ""
    payment_file_type = (data.get("payment_file_type") or "").strip()
    proof_url = None
    proof_note = None
    if payment_file_type == "external":
        stored_name = _extract_upload_filename(payment_file_id)
        proof_url = _build_upload_url(stored_name) if stored_name else None
    else:
        proof_note = "Payment proof was sent in Telegram."
    attendee_rows = db.list_attendees(int(data.get("reservation_id") or 0))
    attendees = [row_a["full_name"] for row_a in attendee_rows]
    repost_proofs = []
    for row_a in attendee_rows:
        if not row_a["repost_discount_applied"]:
            continue
        if (row_a["repost_proof_file_type"] or "").strip() != "external":
            # Non-external repost proofs were sent in Telegram; skip for the web review.
            continue
        stored_name = _extract_upload_filename(row_a["repost_proof_file_id"] or "")
        if not stored_name:
            continue
        repost_proofs.append(
            {
                "full_name": row_a["full_name"],
                "url": _build_upload_url(stored_name),
            }
        )
    return {
        "reservation_id": data.get("reservation_id"),
        "code": data.get("reservation_code"),
        "status": data.get("reservation_status"),
        "quantity": data.get("quantity"),
        "boys": data.get("boys"),
        "girls": data.get("girls"),
        "total_price": data.get("total_price"),
        "base_total_price": data.get("base_total_price"),
        "group_discount_amount": data.get("group_discount_amount"),
        "discount_amount": data.get("discount_amount"),
        "created_at": data.get("created_at"),
        "event_id": data.get("event_id"),
        "event_title": data.get("event_title"),
        "event_datetime": data.get("event_datetime"),
        "buyer_tg_id": data.get("buyer_tg_id"),
        "buyer_name": data.get("buyer_name"),
        "buyer_surname": data.get("buyer_surname"),
        "buyer_phone": data.get("buyer_phone"),
        "attendees": attendees,
        "payment_file_type": payment_file_type,
        "proof_url": proof_url,
        "proof_note": proof_note,
        "repost_proofs": repost_proofs,
    }


@app.get("/api/admin/reservation/pending")
def admin_reservation_pending(request: Request, tg_id: Optional[int] = None) -> Dict[str, Any]:
    _request_admin(request, tg_id)
    rows = db.list_pending_reservations(limit=100)
    return {"items": [_pending_reservation_item(row) for row in rows]}


@app.post("/api/admin/reservation/approve")
def admin_reservation_approve(request: Request, payload: AdminReservationApproveRequest) -> Dict[str, Any]:
    admin_tg_id = _request_admin(request, payload.tg_id)
    ok, message, reservation = db.approve_reservation(payload.reservation_id, admin_tg_id)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    if reservation is not None:
        buyer = db.get_user_by_id(reservation.user_id)
        buyer_tg_id = int(buyer.tg_id) if buyer else 0
        if buyer_tg_id > 0:
            _bot_api(
                "sendMessage",
                {
                    "chat_id": buyer_tg_id,
                    "text": (
                        "Your booking is approved!\n"
                        f"Code: {reservation.code}\n\n"
                        "Your tickets are confirmed. See you at the event!"
                    ),
                },
            )
        buyer_email = (getattr(buyer, "email", "") or "").strip() if buyer else ""
        if buyer_email:
            event = db.get_event(reservation.event_id)
            event_title = event.title if event else f"Event #{reservation.event_id}"
            _send_email(
                buyer_email,
                "Your Budapest Tunderi booking is approved",
                (
                    "Good news! Your booking is approved.\n"
                    f"Event: {event_title}\n"
                    f"Code: {reservation.code}\n\n"
                    "Your tickets are confirmed. Open the website to view your QR passes. "
                    "See you at the event!"
                ),
            )
        if (reservation.payment_file_type or "").strip() == "external":
            _delete_stored_upload(reservation.payment_file_id)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/reservation/reject")
def admin_reservation_reject(request: Request, payload: AdminReservationRejectRequest) -> Dict[str, Any]:
    admin_tg_id = _request_admin(request, payload.tg_id)
    note = (payload.note or "").strip()
    if not note:
        raise HTTPException(status_code=400, detail="Rejection note is required.")
    ok, message, reservation = db.reject_reservation(payload.reservation_id, admin_tg_id, note)
    if not ok:
        raise HTTPException(status_code=400, detail=message)
    if reservation is not None:
        buyer = db.get_user_by_id(reservation.user_id)
        buyer_tg_id = int(buyer.tg_id) if buyer else 0
        if buyer_tg_id > 0:
            _bot_api(
                "sendMessage",
                {
                    "chat_id": buyer_tg_id,
                    "text": (
                        "Your booking was rejected.\n"
                        f"Code: {reservation.code}\n\n"
                        f"Reason: {note}\n\n"
                        "Contact us on Telegram: @budapest_tunderi"
                    ),
                },
            )
        buyer_email = (getattr(buyer, "email", "") or "").strip() if buyer else ""
        if buyer_email:
            event = db.get_event(reservation.event_id)
            event_title = event.title if event else f"Event #{reservation.event_id}"
            _send_email(
                buyer_email,
                "Your Budapest Tunderi booking was rejected",
                (
                    "Unfortunately your booking was rejected.\n"
                    f"Event: {event_title}\n"
                    f"Code: {reservation.code}\n\n"
                    f"Reason: {note}\n\n"
                    "If you have questions, contact us on Telegram: @budapest_tunderi"
                ),
            )
        if (reservation.payment_file_type or "").strip() == "external":
            _delete_stored_upload(reservation.payment_file_id)
    return {"ok": True, "message": message, "reservation": reservation.__dict__ if reservation else None}


@app.post("/api/admin/logout")
def admin_logout(request: Request, response: Response) -> Dict[str, Any]:
    token = _request_admin_token(request)
    if token:
        db.delete_admin_web_session(token)
    response.delete_cookie(ADMIN_SESSION_COOKIE, path="/")
    return {"ok": True}


if __name__ == "__main__":
    import uvicorn

    host = os.getenv("MINI_APP_HOST", "0.0.0.0")
    port = int(os.getenv("PORT") or os.getenv("MINI_APP_PORT", "8080"))
    reload_enabled = os.getenv("MINI_APP_RELOAD", "0") == "1"
    uvicorn.run("ticketbot.miniapp_server:app", host=host, port=port, reload=reload_enabled)
